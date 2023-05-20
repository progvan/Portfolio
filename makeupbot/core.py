from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook
import time
from datetime import datetime
import re
import os

import config
import answer

#перевірка чи присутній excel-файл, якщо присутній - завантажує його
#якщо відсутній - створює excel-файл з трьома листами:
# sheet (основний лист) - лист для графіку,
# sheet_info_users - псевдо база даних для зберігання змінних користувачів
# sheet_logs - лист для зберігання логів
def load_book_file():
    if os.path.isfile(config.file_xlsx_name) == False:
        config.book = Workbook()
        config.sheet = config.book.active
        config.sheet.title = "Shedule"
        config.sheet.column_dimensions[get_column_letter(3)].width = 15
        config.sheet_info_users = config.book.create_sheet(title="Users Info")
        config.sheet_info_users.column_dimensions[get_column_letter(1)].width = 15
        config.sheet_info_users.column_dimensions[get_column_letter(2)].width = 15
        config.sheet_logs = config.book.create_sheet(title="Logs")
        config.sheet_logs['A1'] = "Дата(dmy)"
        config.sheet_logs.column_dimensions[get_column_letter(1)].width = 10
        config.sheet_logs['B1'] = "Час"
        config.sheet_logs.column_dimensions[get_column_letter(2)].width = 13
        config.sheet_logs['C1'] = "Telegram ID"
        config.sheet_logs.column_dimensions[get_column_letter(3)].width = 12
        config.sheet_logs['D1'] = "BOT/USER"
        config.sheet_logs.column_dimensions[get_column_letter(4)].width = 20
        config.sheet_logs['E1'] = "Текст бота або користувача"
        config.sheet_logs.column_dimensions[get_column_letter(5)].width = 70
        save_book()
    config.book = load_workbook(config.file_xlsx_name)
    config.sheet = config.book.active

#додавання id користувачів у базу даних(окремий excel-лист)
#якщо користувач вже присутній - return True
#таблиця пуста - додає в першу комірку - return True
def add_id_users_excel(key):
    config.sheet_info_users = config.book['Users Info']
    if config.sheet_info_users['A1'].value == None:
        config.sheet_info_users['A1'].value = key
        save_book()
        return True
    else:
        for row in range(1, config.sheet_info_users.max_row + 1):
            if config.sheet_info_users[row][0].value == key:
                return True
        config.sheet_info_users[config.sheet_info_users.max_row + 1][0].value = key
        save_book()

#орієнтир, повертає значення на якому місці стоїть клас користувача у списку через його id
def get_user_key(user_id):
    key = 0
    for user in config.ListUsersID:
        if user.UserId == user_id:
            return key
        else:
            key += 1

#перевірка чи є користувач в системі
def check_user_in_system(user_id):
    for user in config.ListUsersID:
        if user.UserId == user_id:
            return True
    return False

#повертає користувачеві запис через його айді
def pull_user_record(user_id):
    config.sheet_info_users = config.book['Users Info']
    for row in range(1, config.sheet_info_users.max_row + 1):
        if int(config.sheet_info_users[row][0].value) == int(user_id):
            return config.sheet_info_users[row][1].value

#функція слідкування змінної lock у кожного користувача
def tracking_flag_lock(key_user):
    config.sheet_info_users = config.book['Users Info']
    for row in range(1, config.sheet_info_users.max_row + 1):
        if row == key_user + 1:
            if config.sheet_info_users['B' + str(row)].value != None:
                if time.strptime(config.sheet_info_users[row][1].value, '%d.%m %H:%M') < time.strptime(init_today(), '%d.%m %H:%M'):
                    config.sheet_info_users['B' + str(row)].value =  None
                    config.sheet_info_users['C' + str(row)].value =  None
                    clearing_variables(key_user)
                    config.ListUsersID[key_user].lock = 0
                else:
                    clearing_variables(key_user)
                    config.ListUsersID[key_user].lock = 2
                    config.sheet_info_users.cell(row, 3, "lock")
                save_book()
            elif config.ListUsersID[key_user].record_type != "":
                config.ListUsersID[key_user].lock = 3
            else:
                clearing_variables(key_user)
                config.ListUsersID[key_user].lock = 0
            break

#перевіряє чи є надана дата в таблиці, якщо є - повертає 0, немає - 1
def check_date(string_date, leng):
    for counter in range(leng):
        for row1 in range(1, config.sheet.max_row + 1):
            if string_date[counter].split(" ")[0] == config.sheet[row1][0].value:
                if string_date[counter].split(" ")[1] == config.sheet[row1][1].value:
                    return 0
    return 1

#додавання класів у список з таблиці через id
def init_userList(user_id):
    if check_user_in_system(user_id):
        return True
    else:
        config.ListUsersID = []
        for row in range(1, config.sheet_info_users.max_row + 1):
            config.ListUsersID.append(config.Users(config.sheet_info_users[row][0].value))
            if config.sheet_info_users['C' + str(row)].value != None:
                config.ListUsersID[row - 1].lock = 2

#додавання запису у лист інфи про користувача та зберігання її
def add_record_user_excel(record_user, key_user):
    for row in range(1, config.sheet_info_users.max_row + 1):
        if row == key_user + 1:
            config.sheet_info_users.cell(row, 2, record_user)
            break
    save_book()

#конструктор класа
def clearing_variables(key_user):
    config.ListUsersID[key_user].first_name = ""
    config.ListUsersID[key_user].last_name = ""
    config.ListUsersID[key_user].contact = 0
    config.ListUsersID[key_user].message_id_contact = 0
    config.ListUsersID[key_user].instagram_username = ""
    config.ListUsersID[key_user].record_type = ""
    config.ListUsersID[key_user].record_user = ""
    config.ListUsersID[key_user].comment_user = ""

#очищення запису користувача
def delete_record_sheet_users(user_id):
    config.sheet_info_users = config.book['Users Info']
    for row in range(1, config.sheet_info_users.max_row + 1):
        if int(config.sheet_info_users[row][0].value) == int(user_id):
            config.sheet_info_users['B' + str(row)].value =  None
            config.sheet_info_users['C' + str(row)].value =  None
            clearing_variables(row-1)
            config.ListUsersID[row - 1].lock = 1
            break

#запис логів
def add_LOG(user_id, text = "bot text", first_name = "bot", last_name = "reply"):
    config.sheet_logs = config.book['Logs']
    config.sheet_logs[config.sheet_logs.max_row + 1][0].value = str(datetime.now().date().strftime('%d-%m-%y'))
    config.sheet_logs[config.sheet_logs.max_row][1].value = str(datetime.now().time().strftime('%H:%M:%S.%f')[:-4])
    if user_id == config.admin_id:
        config.sheet_logs[config.sheet_logs.max_row][2].value = "ADMIN"
        config.sheet_logs[config.sheet_logs.max_row][3].value = "ADMIN"
    else:
        config.sheet_logs[config.sheet_logs.max_row][2].value = user_id
        config.sheet_logs[config.sheet_logs.max_row][3].value = first_name + " " + last_name
    config.sheet_logs[config.sheet_logs.max_row][4].value = text
    save_book()

#повертає сьогоднішні день.місяць години:хвилини
#ЗАУВАЖУ, години + 2, тому що на сайті на якому бот хостится години на 2 раніше чим укр
def init_today():
    return str(datetime.now().today().day) + "." + str(datetime.now().today().month) + " " + str(datetime.now().today().hour) + ":" + str(datetime.now().today().minute)

#розділяє строку на дату та час і записує це в таблицю
def data_recording(string_date):
    if config.sheet['A1'].value == None:
        config.sheet['A1'].value = string_date.split()[0]
        config.sheet['B1'].value = string_date.split()[1]
    else:
        config.sheet[config.sheet.max_row + 1][0].value = string_date.split()[0]
        config.sheet[config.sheet.max_row][1].value = string_date.split()[1]

#сортує рядки
def sort_col():
    column2 = []
    for row in range(1, config.sheet.max_row + 1):
        column1 = []
        column1.append(time.strptime(str(config.sheet[row][0].value), "%d.%m"))
        column1.append(time.strptime(str(config.sheet[row][1].value), "%H:%M"))
        if config.sheet['C' + str(row)].value != None:
                column1.append(str(config.sheet['C' + str(row)].value))
                config.sheet['C' + str(row)] = None
        column2.append(column1)
    column2.sort(key = lambda x: x[:2])
    for row in range(1, config.sheet.max_row + 1):
        config.sheet[row][0].value = time.strftime("%d.%m", column2[row-1][0])
        config.sheet[row][1].value = time.strftime("%H:%M", column2[row-1][1])
        if len(column2[row - 1]) == 3:
                config.sheet[row][2].value = column2[row - 1][2]

#видаляє застарілі записи
def del_old_data():
    rowsList = []
    for row in range(1, config.sheet.max_row + 1):
        config.sheet = config.book['Shedule']
        if time.strptime(str(config.sheet[row][0].value) + " " + str(config.sheet[row][1].value), "%d.%m %H:%M") <= time.strptime(init_today(), "%d.%m %H:%M"):
            rowsList.append(row)
            if config.sheet['C' + str(row)].value != None:
                add_LOG(config.sheet[row][2].value, "delete old data")
                delete_record_sheet_users(config.sheet[row][2].value)
    config.sheet = config.book['Shedule']
    for i in reversed(rowsList):
        config.sheet.delete_rows(i)

#функція адміну
#виводить усі записи, якщо запис кимось зайнятий, то виводить поряд з ним що він зайнятий
def output_sort_data_admin():
    config.sheet = config.book['Shedule']
    if config.sheet['A1'].value == None:
        save_book()
        return answer.TEXT_EMPTY_SHEDULE__ADMIN
    sort_col()
    del_old_data()
    save_book()
    if config.sheet['A1'].value == None:
        return answer.TEXT_EMPTY_SHEDULE__ADMIN
    bot_mes1 = ""
    for row1 in range(1, config.sheet.max_row + 1):
        bot_mes1 += str(row1) + ". " + str(config.sheet[row1][0].value) + " - " + str(config.sheet[row1][1].value)
        if config.sheet['C' + str(row1)].value != None:
            bot_mes1 += " (зайнято)"
        bot_mes1 += "\n"
    return bot_mes1

#функція користувача
#виводить тільки вільні записи, тобто ті, що не ким не зайняті
def output_for_users():
    if config.sheet['A1'].value == None:
        return False
    del_old_data()
    save_book()
    bot_mes1 = ""
    counter = 1
    for row1 in range(1, config.sheet.max_row + 1):
        if config.sheet['C' + str(row1)].value == None:
            bot_mes1 += str(counter) + ". " + str(config.sheet[row1][0].value) + " - " + str(config.sheet[row1][1].value) + "\n"
            counter += 1
    if bot_mes1 == "":
        return False
    else:
        return bot_mes1

#записує користувача в момент коли він обрав свій запис та надіслав свій рядок
def enroll_user(value_user, key_user):
    counter = 1
    temp = value_user.split(' ')
    for row1 in range(1, config.sheet.max_row + 1):
        if config.sheet['C' + str(row1)].value == None:
            if temp[0] == config.sheet[row1][0].value:
                if temp[1] == config.sheet[row1][1].value:
                    config.sheet['C' + str(row1)].value = config.ListUsersID[key_user].UserId
                    break
            counter += 1
    save_book()

#повертає скільки всього вільних місць (потрібно для контролювання чи корректні номери вводить користувач)
def value_free_places_for_users():
    x = 0
    for i in range(1, config.sheet.max_row + 1):
        if config.sheet['C' + str(i)].value != None:
            x += 1
    return config.sheet.max_row - x

#повертає запис у форматі ХХ.ХХ ХХ:ХХ який обраний користувачем на даний момент
def info_check_user(value_user):
    return output_for_users().split('\n')[value_user-1][3:].replace(' -', '', 1)

#перевірка чи є слові кирилиця, пробіли та @ (потрібно для перевірки коректності інстаграм ніка)
def check_text_instagram_username(instagram_username):
    for x in instagram_username:
        if re.search(r'[а-яА-ЯёЁіІїЇєЄґҐ\s@]', x):
            return False
    return True

#повертає рядок з інформацією для користувача
def info_record_user(user_id):
    return ("Ви записані на " + str(config.ListUsersID[get_user_key(user_id)].record_user)
        + "\nТип процедури: " + str(config.ListUsersID[get_user_key(user_id)].record_type)
        + "\nInstagram: " + str(config.ListUsersID[get_user_key(user_id)].instagram_username)
        + "\nВаш коментар: " + str(config.ListUsersID[get_user_key(user_id)].comment_user))

#повертає рядок з інформацією для адміна
def info_record_admin(user_id):
    return ("Запис на " + str(config.ListUsersID[get_user_key(user_id)].record_user)
        + "\nТип процедури: " + str(config.ListUsersID[get_user_key(user_id)].record_type)
        + "\n\n--Інформація про клієнта--"
        + "\nІм'я та прізвище (у Telegram): " + str(config.ListUsersID[get_user_key(user_id)].first_name) + " " + str(config.ListUsersID[get_user_key(user_id)].last_name)
        + "\nТелефон: " + str(config.ListUsersID[get_user_key(user_id)].contact)
        + "\nInstagram: " + str(config.ListUsersID[get_user_key(user_id)].instagram_username)
        + "\nКоментар: " + str(config.ListUsersID[get_user_key(user_id)].comment_user))

#зберігає книгу
def save_book():
    config.book.save(config.file_xlsx_name)
    config.book.close()

